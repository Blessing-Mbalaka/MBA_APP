from django.http import HttpResponseNotFound
from django.shortcuts import redirect, render,get_object_or_404
from django.urls import reverse
from mbamain.models import AUser, Invite, StudentProfile, Project
from mbaAdmin.utils import is_admin, send_invite,valid_role_type,generate_temp_password, send_invite_email, is_valid_email, supervisor_matches_discipline
from django.contrib import messages
from django.db.models import Q
from openpyxl import load_workbook
import threading
from django.db import transaction




@is_admin
def students(request):
    per_page = 10
    page = request.GET.get("page", 0)
    search = request.GET.get("search",'')

    is_searching = True if search and  (not search == '') else False
    student_no = request.GET.get("student_no")
    block_id = request.GET.get("block_id",'')
    query = 'block_id={block_id}&student_no={student_no}&search={search}' if is_searching else '' 

    try:
        page = int(page)
        if page < 0:
            page = 0
    except:
        page = 0
        
        
    has_next = True
    next_page = page + 1
    prev_page = page - 1 if page > 0 else 0

    next_url = reverse("mba_admin:students") + f"?page={next_page}&{query}"
    prev_url = reverse("mba_admin:students") + f"?page={prev_page}&{query}"
    if is_searching:
        filters = (
        Q(student_profile__student_no=student_no) &
        Q(user_type = AUser.UserType.STUDENT)
       )
        students = AUser.objects.filter(filters).order_by('-date_joined')[page*per_page:(page+1)*per_page]
    else :
        students = AUser.objects.filter(Q(user_type = AUser.UserType.STUDENT)).order_by('-date_joined')[page*per_page:(page+1)*per_page]
        print(page, per_page, students.count())
    if len(students) == 0 and page > 0: 
        page = page - 1
        next_url = reverse("mba_admin:students") + f"?page={page}&{query}"
        return redirect(next_url)
    return render(request, "mbaAdmin/students.html", {"students": students, 
                                                      "next":next_url, 
                                                      "prev": prev_url , 
                                                      "has_next": has_next, 
                                                      "has_prev": True if page > 0 else False,  
                                                       "student_no": student_no,
                                                       "block_id": block_id
                                                        })
    


@is_admin
def student_bulk_onboard(request):
    if request.method == 'POST':
        file = request.FILES.get('file')
        block_id = request.POST.get('block_id')

        if not block_id or block_id == 'None':
            messages.error(request, "Please select a valid block ID.")
            return redirect(reverse("mba_admin:students"))

        if not file:
            messages.error(request, "Please upload a file.")
            return redirect(reverse("mba_admin:students"))

        if not file.name.endswith('.xlsx'):
            messages.error(request, "Please upload a valid excel file.")
            return redirect(reverse("mba_admin:students"))

        try:
            workbook = load_workbook(file)
            ws = workbook.active
            failed_count = 0
            failed_emails = ''
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Validate row has enough columns (7 required)
                # Template2: Title, Last name, First name, Contact, Student Number, Email address, Secondary
                if not row or len(row) < 7:
                    failed_emails += f" [Row {row_num}: Missing columns - need 7]"
                    failed_count += 1
                    continue
                    
                try:
                    # Unpack columns: Title, Last name, First name, Contact, Student Number, Email address, Secondary
                    title = row[0]
                    surname = row[1]  # "Last name"
                    name = row[2]     # "First name"
                    contact = row[3]
                    student_no = row[4]
                    student_email = row[5]
                    secondary_email = row[6]

                    # Validate required fields
                    if not student_email or not name or not surname or not student_no:
                        failed_emails += f" [Row {row_num}: Missing required fields]"
                        failed_count += 1
                        continue
                    
                    # Normalize email and student_no for comparison
                    student_email_clean = str(student_email).strip()
                    student_no_clean = str(student_no).strip()
                    
                    # Better duplicate checking - check both user AND profile (case-insensitive)
                    user_exists = AUser.objects.filter(email__iexact=student_email_clean).exists()
                    profile_exists = StudentProfile.objects.filter(
                        Q(student_no__iexact=student_no_clean) | Q(user__email__iexact=student_email_clean)
                    ).exists()
                    
                    if user_exists or profile_exists:
                        failed_count += 1
                        continue

                    if not is_valid_email(student_email_clean):
                        
                        failed_emails += " " + student_email_clean
                        continue

                    temp_pass = generate_temp_password()

                    with transaction.atomic():
                        # Create user (signal will auto-create StudentProfile)
                        user = AUser.objects.create_user(
                            username=student_email_clean,
                            email=student_email_clean,
                            password=temp_pass,
                            user_type=AUser.UserType.STUDENT
                        )

                        # Get the auto-created profile from signal and update it with data
                        student_profile, _ = StudentProfile.objects.get_or_create(user=user)
                        student_profile.name = name
                        student_profile.surname = surname
                        student_profile.title = title
                        student_profile.contact = contact
                        student_profile.student_no = student_no_clean
                        student_profile.secondary_email = secondary_email
                        student_profile.module = "NA"
                        student_profile.block_id = block_id
                        student_profile.save()

                        # Mark user as having profile
                        user.has_profile = True
                        user.save()

                        # Only send invite if commit succeeds
                        transaction.on_commit(
                            lambda u=user, pwd=temp_pass: threading.Thread(
                                target=send_invite, args=(u, pwd)
                            ).start()
                        )

                except Exception as e:
                    failed_emails += f" [Row {row_num}: {str(e)}]"
                    print(f"Error at row {row_num}: {str(e)}")
                    failed_count += 1
                    continue

            messages.success(
                request,
                f"Students uploaded successfully. Failed/skipped: {failed_emails if failed_emails else 'None'}"
            )
            # if failed_count > 0:
            #     messages.warning(request, f"{failed_count} students already existed or had invalid data.")

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            print(e)

        return redirect(reverse("mba_admin:students"))

    return HttpResponseNotFound("Page not found.")

@is_admin
def manage_student(request, id):
    student = get_object_or_404(AUser, pk=id)

    try:
        project = student.projects.all()[0]
        supervisor = project.get_supervisor()
        invites = project.projects_invites.filter(read=False, invite_type=False)
    except IndexError:
        project = None
        supervisor = None
        invites = None
  
    if project :
        # Get all scholars and apply smart discipline matching
        all_supervisors = AUser.objects.filter(user_type=AUser.UserType.SCHOLAR).order_by('supervisor_profile__students')
        supervisors = [
            sup for sup in all_supervisors 
            if supervisor_matches_discipline(sup.supervisor_profile.skills or '', project.discipline)
        ][0:10]
        responses_invites = Invite.objects.filter(response=1, project=project, invite_type=False)
        sup = [supervisor for supervisor in supervisors if not Invite.objects.filter(project=project, user=supervisor)]
    else:
        sup = []
        responses_invites= []
        
    return render(request, 'mbaAdmin/manageStudent.html', {'student': student, "supervisors": sup, "project": project, "responses": responses_invites, "supervisor": supervisor, "invites": invites, "size": len(sup)})


@is_admin
def reset_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    project.reset_project()
    messages.success(request, "Project has been reseted successfully")
    project.projects_invites.all().delete()
    return redirect("mba_admin:manage_student", project.student.id)