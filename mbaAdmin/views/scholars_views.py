from django.http import HttpResponseNotFound
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from mbamain.models import AUser, SupervisorProfile, ExamminerProfile, Project
from mbaAdmin.utils import is_admin, send_invite,valid_role_type,generate_temp_password, is_valid_email
from django.contrib import messages
from django.db.models import Q
from openpyxl import load_workbook
import threading 
from django.db import transaction

@is_admin
def scholars(request):
    per_page = 10
    page = request.GET.get("page", 0)
    search = request.GET.get("search",'')

    is_searching = True if search and  (not search == '') else False
    experience = request.GET.get("experience")
    name = request.GET.get("name", '')
    title = request.GET.get("title")
    position  = request.GET.get("position")
    role  = request.GET.get("role", AUser.RoleType.BOTH)
    query = 'name={experience}&title={title}&position={position}&role={role}' if is_searching else '' 
    try:
        page = int(page)
        if page < 0:
            page = 0
    except:
        page = 0
        
    has_next = True
    next_page = page + 1
    prev_page = page - 1 if page > 0 else 0

    next_url = reverse("mba_admin:scholars") + f"?page={next_page}&{query}"
    prev_url = reverse("mba_admin:scholars") + f"?page={prev_page}&{query}"
    if is_searching:
        filters = (
    ( Q(profile__name__icontains=name) | Q(profile__surname__icontains=name) |
    Q(role_type=role) |
    Q(profile__experience=experience) |
    Q(profile__title=title) |
    Q(profile__position=position)) & Q(has_profile=True)
)
        scholars = AUser.objects.filter(filters).order_by('date_joined')[page*per_page:(page+1)*per_page]
    else :
        scholars = AUser.objects.filter((Q(role_type = AUser.RoleType.BOTH) | Q(role_type = AUser.RoleType.EXAMINER) | Q(role_type = AUser.RoleType.SUPERVISOR)) & Q(has_profile=True) ).order_by('date_joined')[page*per_page:(page+1)*per_page]
    
    if len(scholars) == 0 and page > 0: # no more scholars on this page, redirect to the previous page
        page = page - 1
        next_url = reverse("mba_admin:scholars") + f"?page={page}&{query}"
        return redirect(next_url)
    return render(request, "mbaAdmin/scholars.html", {"scholars": scholars, 
                                                      "next":next_url, 
                                                      "prev": prev_url , 
                                                      "has_next": has_next, 
                                                      "has_prev": True if page > 0 else False,
                                                       "role": role,
                                                       "experience": experience,
                                                        "title": title,
                                                        "position": position,
                                                        "name": name
                                                      })

@is_admin
def supervisors(request):
    per_page = 10
    page = request.GET.get("page", 0)
    search = request.GET.get("search",'')

    is_searching = True if search and  (not search == '') else False
    name = request.GET.get("name", '')
    query = 'name={name}' if is_searching else '' 
    try:
        page = int(page)
        if page < 0:
            page = 0
    except:
        page = 0
        
    has_next = True
    next_page = page + 1
    prev_page = page - 1 if page > 0 else 0
     
    next_url = reverse("mba_admin:supervisors") + f"?page={next_page}&{query}"
    prev_url = reverse("mba_admin:supervisors") + f"?page={prev_page}&{query}"
    if is_searching:
        filters = (
     (Q(supervisor_profile__name__icontains=name) | 
     Q(supervisor_profile__surname__icontains=name) | 
     Q(supervisor_profile__name__icontains=name) |
     Q(email__icontains=name)) & Q(user_type = AUser.UserType.SCHOLAR) 
    
)
        scholars = AUser.objects.filter(filters).order_by('date_joined')[page*per_page:(page+1)*per_page]
    else :
        scholars = AUser.objects.filter((Q(user_type = AUser.UserType.SCHOLAR)  | Q(role_type = AUser.RoleType.SUPERVISOR)) & Q(has_profile=True) ).order_by('date_joined')[page*per_page:(page+1)*per_page]
    
    if len(scholars) == 0 and page > 0: 
        page = page - 1
        next_url = reverse("mba_admin:supervisors") + f"?page={page}&{query}"
        return redirect(next_url)
    return render(request, "mbaAdmin/supervisors.html", {"scholars": scholars, 
                                                      "next":next_url, 
                                                      "prev": prev_url , 
                                                      "has_next": has_next, 
                                                      "has_prev": True if page > 0 else False,
                                                        "name": name
                                                      })


@is_admin
def examiners(request):
    per_page = 10
    page = request.GET.get("page", 0)
    search = request.GET.get("search",'')

    is_searching = True if search and  (not search == '') else False
    name = request.GET.get("name", '')
    query = 'name={name}' if is_searching else '' 
    try:
        page = int(page)
        if page < 0:
            page = 0
    except:
        page = 0
        
    has_next = True
    next_page = page + 1
    prev_page = page - 1 if page > 0 else 0

    next_url = reverse("mba_admin:examiners") + f"?page={next_page}&{query}"
    prev_url = reverse("mba_admin:examiners") + f"?page={prev_page}&{query}"
    if is_searching:
        filters = (
     (Q(name__icontains=name) | Q(surname__icontains=name) |
    Q(email__icontains=name))  
)
        examiners = ExamminerProfile.objects.filter(filters).order_by('-created_at')[page*per_page:(page+1)*per_page]
    else :
        examiners = ExamminerProfile.objects.filter().order_by('-created_at')[page*per_page:(page+1)*per_page]
    
    if len(examiners) == 0 and page > 0: 
        page = page - 1
        next_url = reverse("mba_admin:examiners") + f"?page={page}&{query}"
        return redirect(next_url)
    return render(request, "mbaAdmin/examiners.html", {"examiners": examiners, 
                                                      "next":next_url, 
                                                      "prev": prev_url , 
                                                      "has_next": has_next, 
                                                      "has_prev": True if page > 0 else False,
                                                        "name": name
                                                      })



@is_admin
def onboard_examiners_bulk(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Please upload a file.")
            return redirect("mba_admin:examiners")

        if not file.name.endswith('.xlsx'):
            messages.error(request, "Please upload a valid excel file.")
            return redirect("mba_admin:examiners")

        try:
            wb = load_workbook(file)
            ws = wb.active
            existence_count = 0
            failed = ""

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Validate row has enough columns (13 required)
                if not row or len(row) < 13:
                    failed += f" [Row {row_num}: Missing columns - need 13]"
                    continue

                try:
                    # Unpack columns: Name, Surname, Title, Qualification, Affiliation, Street_address,
                    # Cell_phone, Email, Number_of_students_supervised, Current_affiliation,
                    # Number_publications, International_assessor, Academic_experience
                    name = row[0]
                    surname = row[1]
                    title = row[2]
                    qualification = row[3]
                    affiliation = row[4]
                    street_address = row[5]
                    cell_phone = row[6]
                    email = row[7]
                    number_of_students_supervised = row[8] if row[8] is not None else 0
                    current_affiliation = row[9]
                    number_publications = row[10] if row[10] is not None else 0
                    international_assessor = row[11] if row[11] is not None else False
                    academic_experience = row[12] if row[12] is not None else 0

                    # Validate required fields
                    if not email or not name or not surname:
                        failed += f" [Row {row_num}: Missing Name, Surname, or Email]"
                        continue

                    if not is_valid_email(str(email).strip()):
                        failed += " " + str(email)
                        continue

                    if ExamminerProfile.objects.filter(email=str(email).strip()).exists():
                        existence_count += 1
                        continue

                    temp_password = generate_temp_password()

                    with transaction.atomic():
                        # Create user
                        user = AUser.objects.create_user(
                            username=str(email).strip(),
                            email=str(email).strip(),
                            password=temp_password,
                        )
                        user.set_role_type(AUser.RoleType.EXAMINER)
                        user.set_user_type(AUser.UserType.EXAMINER)

                        # Create profile
                        profile = ExamminerProfile.objects.create(
                            user=user,
                            name=name,
                            surname=surname,
                            title=title,
                            qualification=qualification,
                            affiliation=affiliation,
                            street_address=street_address,
                            cell_phone=cell_phone,
                            email=str(email).strip(),
                            number_of_students_supervised=int(number_of_students_supervised) if number_of_students_supervised else 0,
                            current_affiliation=current_affiliation,
                            number_publications=int(number_publications) if number_publications else 0,
                            international_assessor=bool(international_assessor),
                            academic_experience=int(academic_experience) if academic_experience else 0
                        )

                        # Mark user as having profile
                        user.has_profile = True
                        user.save()

                        # Schedule invite after commit
                        transaction.on_commit(
                            lambda u=user, pwd=temp_password: threading.Thread(
                                target=send_invite, args=(u, pwd)
                            ).start()
                        )

                except Exception as e:
                    failed += f" [Row {row_num}: {str(e)}]"
                    print(f'Error at row {row_num}: {str(e)}')
                    print(f"Failed to onboard examiner: {email if 'email' in locals() else 'unknown'}, {name if 'name' in locals() else 'unknown'}")
                    continue

            messages.success(
                request,
                f"Examiners uploaded successfully. Failed/skipped: {failed if failed else 'None'}"
            )

        except Exception as e:
            print(e)
            messages.error(request, f"An error occurred while processing the file: {str(e)}")

        if existence_count > 0:
            messages.warning(
                request,
                f"{existence_count} examiners already existed. They were not created again."
            )

        return redirect("mba_admin:examiners")

    return HttpResponseNotFound("Page not found")

@is_admin
def onboard_scholar(request):
    pass
    return  HttpResponseNotFound("Page not found")




@is_admin
def onboard_supervisor_bulk(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Please upload a file.")
            return redirect("mba_admin:supervisors")

        if not file.name.endswith('.xlsx'):
            messages.error(request, "Please upload a valid excel file.")
            return redirect("mba_admin:supervisors")

        try:
            wb = load_workbook(file)
            ws = wb.active
            existence_count = 0
            failed = ''

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Validate row has enough columns (5 required)
                # Template3: Title, Names, Surname, Contact Details, email
                if not row or len(row) < 5:
                    failed += f" [Row {row_num}: Missing columns - need 5]"
                    continue

                try:
                    # Unpack columns: Title, Names, Surname, Contact Details, email
                    title = row[0]
                    names = row[1]  # "Names"
                    surname = row[2]
                    contact = row[3]  # "Contact Details"
                    email = row[4]

                    # Validate required fields
                    if not email or not names or not surname:
                        failed += f" [Row {row_num}: Missing required fields]"
                        continue

                    if not is_valid_email(str(email).strip()):
                        failed += " " + str(email)
                        print('invalid email')
                        continue

                    if AUser.objects.filter(Q(email=str(email).strip()) | Q(username=str(email).strip())).exists():
                        existence_count += 1
                        user = AUser.objects.get(email=str(email).strip())
                        print(user.supervisor_profile)
                        continue

                    temp_password = generate_temp_password()

                    with transaction.atomic():
                        # Create user
                        user = AUser.objects.create_user(
                            username=str(email).strip(),
                            email=str(email).strip(),
                            password=temp_password,
                        )
                        user.set_role_type(AUser.RoleType.SUPERVISOR)
                        user.set_user_type(AUser.UserType.SCHOLAR)

                        # Create profile
                        profile = SupervisorProfile.objects.create(
                            user=user,
                            name=names,
                            surname=surname,
                            title=title,
                            contact=contact,
                        )

                        # Mark user as having profile
                        user.has_profile = True
                        user.save()

                        # Schedule invite *only after commit succeeds*
                        transaction.on_commit(
                            lambda u=user, pwd=temp_password: threading.Thread(
                                target=send_invite, args=(u, pwd)
                            ).start()
                        )

                except Exception as e:
                    failed += f" [Row {row_num}: {str(e)}]"
                    print(f"Error at row {row_num}: {str(e)}")
                    print(f"Failed to onboard supervisor: {email if 'email' in locals() else 'unknown'}, {names if 'names' in locals() else 'unknown'}")
                    continue

            messages.success(
                request,
                f"Supervisors uploaded successfully. Failed/skipped: {failed if failed else 'None'}"
            )

        except Exception as e:
            messages.error(request, f"An error occurred while processing the file: {str(e)}")
            print(str(e))

        if existence_count > 0:
            messages.warning(
                request,
                f"{existence_count} supervisors already existed. They were not created again."
            )

        return redirect("mba_admin:supervisors")

    return HttpResponseNotFound("Page not found")



@is_admin
def appoint_assessor(request):
    if request.method == "GET":
        project_id = request.GET.get("project")
        assessor_id = request.GET.get("assessor")
        assessor_no = request.GET.get("assessor_no")
        nomination_no = request.GET.get("nomination_no")
        
        if not assessor_no in ['1','2'] or nomination_no not in ['1','2','3']:
            return HttpResponseNotFound()
      
        project = get_object_or_404(Project, pk=project_id)
        if  (nomination_no == '1' and not project.assessor_1_approved):
            messages.error(request, "You can not appoint an assessor that was not approved by HDC")
            return redirect('mba_admin:nomination_submitted')
        if  (nomination_no == '2' and not project.assessor_2_approved):
            messages.error(request, "You can not appoint an assessor that was not approved by HDC")
            return redirect('mba_admin:nomination_submitted')
        
        if  (nomination_no == '3' and not project.assessor_3_approved):
            messages.error(request, "You can not appoint an assessor that was not approved by HDC")
            return redirect('mba_admin:nomination_submitted')
        

        if nomination_no == '1' and not project.assessor_1_response:
            messages.error(request, "You can not appoint an assessor that has not accepted the invite")
            return redirect('mba_admin:nomination_submitted')
        if nomination_no == '2' and not project.assessor_2_response:
            messages.error(request, "You can not appoint an assessor that has not accepted the invite")
            return redirect('mba_admin:nomination_submitted')
        if nomination_no == '3' and not project.assessor_3_response:
            messages.error(request, "You can not appoint an assessor that has not accepted the invite")
            return redirect('mba_admin:nomination_submitted')
        
        
        if assessor_no == '1':
            assessor = get_object_or_404(ExamminerProfile, pk=assessor_id)
            project.assessor_1 = assessor.id
            project.assessor_1_name = assessor.name
            project.assessor_1_surname = assessor.surname
            project.assessor_1_email = assessor.email
            project.assessor_1_appointed = True
            assessor.save()
        elif assessor_no == '2':
            assessor = get_object_or_404(ExamminerProfile, pk=assessor_id)
            project.assessor_2 = assessor.id
            project.assessor_2_name = assessor.name
            project.assessor_2_surname = assessor.surname
            project.assessor_2_email = assessor.email
            project.assessor_2_appointed = True 
            assessor.save()
        project.save()
        print(project.assessor_1_name)
        print(project.assessor_2_name)
        
        messages.success(request, "Assessor appointed successfuly")
        return redirect('mba_admin:nomination_submitted')
    return HttpResponseNotFound()
    

@is_admin
def update_assessor(request):
    if request.method == "POST":
        project_id = request.POST.get("project")
        print(project_id)
        name = request.POST.get("name")
        surname = request.POST.get("surname")
        email = request.POST.get("email")
        project = get_object_or_404(Project, pk=project_id)
      
        project.assessor_name = name
        project.assessor_surname = surname
        project.assessor_email = email
        project.save()
        messages.success(request, "Assessor updated successfuly")
        return redirect('mba_admin:nomination_submitted')
    return HttpResponseNotFound()
    

