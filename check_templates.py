"""
Check if template files match the code requirements
"""
from openpyxl import load_workbook
import os

templates = {
    'Template1.xlsx': {
        'path': 'mbaAdmin/static/mbaAdmin/templates/Template1.xlsx',
        'name': 'EXAMINER',
        'required_columns': 13,
        'headers': ['Name', 'Surname', 'Title', 'Qualification', 'Affiliation', 'Street_address',
                   'Cell_phone', 'Email', 'Number_of_students_supervised', 'Current_affiliation',
                   'Number_publications', 'International_assessor', 'Academic_experience']
    },
    'Template2.xlsx': {
        'path': 'mbaAdmin/static/mbaAdmin/templates/Template2.xlsx',
        'name': 'STUDENT',
        'required_columns': 7,
        'headers': ['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary']
    },
    'Template3.xlsx': {
        'path': 'mbaAdmin/static/mbaAdmin/templates/Template3.xlsx',
        'name': 'SUPERVISOR',
        'required_columns': 5,
        'headers': ['Title', 'Names', 'Surname', 'Contact Details', 'email']
    }
}

print("\n" + "="*90)
print("TEMPLATE VERIFICATION REPORT")
print("="*90)

all_pass = True

for template_name, spec in templates.items():
    filepath = spec['path']
    
    if not os.path.exists(filepath):
        print(f"\n✗ {spec['name']} TEMPLATE ({template_name})")
        print("-"*90)
        print(f"  ✗ FILE NOT FOUND at {filepath}")
        all_pass = False
        continue
    
    print(f"\n✓ {spec['name']} TEMPLATE ({template_name})")
    print("-"*90)
    print(f"  File: {filepath}")
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for col in range(1, spec['required_columns'] + 1):
            cell_value = ws.cell(1, col).value
            headers.append(cell_value)
        
        print(f"  Columns: {len(headers)}")
        
        # Check column count
        if len(headers) == spec['required_columns']:
            print(f"  ✓ Column count correct: {spec['required_columns']}")
        else:
            print(f"  ✗ Column count MISMATCH: Expected {spec['required_columns']}, got {len(headers)}")
            all_pass = False
        
        # Check header names
        print(f"\n  Expected Headers:")
        for idx, (expected, actual) in enumerate(zip(spec['headers'], headers), 1):
            match = "✓" if expected.lower() == (actual or '').lower() else "✗"
            print(f"    {idx:2d}. {match} Expected: '{expected}'  | Actual: '{actual}'")
        
        # Validate headers match
        headers_match = all(
            expected.lower() == (actual or '').lower() 
            for expected, actual in zip(spec['headers'], headers)
        )
        
        if headers_match:
            print(f"\n  ✓ All headers match required format")
        else:
            print(f"\n  ✗ Headers DO NOT match - template needs correction")
            all_pass = False
            
    except Exception as e:
        print(f"  ✗ Error reading file: {str(e)}")
        all_pass = False

print("\n" + "="*90)
print("SUMMARY")
print("="*90)

if all_pass:
    print("✓✓✓ ALL TEMPLATES ARE CORRECTLY CONFIGURED ✓✓✓")
else:
    print("⚠ SOME TEMPLATES NEED CORRECTION")
    print("\nNeed to update templates to match correct column structure")

print("="*90)
print()
