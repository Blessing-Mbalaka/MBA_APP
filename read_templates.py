from openpyxl import load_workbook
import os

template_files = [
    'mbaAdmin/static/mbaAdmin/templates/Template1.xlsx',
    'mbaAdmin/static/mbaAdmin/templates/Template2.xlsx',
    'mbaAdmin/static/mbaAdmin/templates/Template3.xlsx',
]

for template in template_files:
    try:
        wb = load_workbook(template)
        ws = wb.active
        
        # Get first row (headers)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        print(f"\n{'='*70}")
        print(f"File: {template}")
        print(f"{'='*70}")
        print(f"Total Columns: {len(headers)}")
        print(f"\nColumn Headers:")
        for i, header in enumerate(headers, 1):
            print(f"  {i:2d}. {header}")
        
        # Show first data row as example
        print(f"\nFirst Data Row (Example):")
        first_data_row = []
        for i, cell in enumerate(ws[2]):
            if i < len(headers):
                first_data_row.append(cell.value)
        for i, value in enumerate(first_data_row, 1):
            print(f"  {i:2d}. {value}")
            
    except Exception as e:
        print(f"Error reading {template}: {e}")

print(f"\n{'='*70}")
