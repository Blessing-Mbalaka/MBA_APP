"""
Quick test - verify bulk uploads work (no cleanup in middle)
"""
import os
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import Client
from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile, ExamminerProfile, SupervisorProfile, AUser

User = get_user_model()


def test_uploads():
    # Create admin
    admin, _ = User.objects.get_or_create(
        username='testadmin1',
        defaults={'email': 'testadmin1@test.com', 'user_type': User.UserType.ADMIN}
    )
    admin.set_password('pass')
    admin.save()
    
    client = Client()
    client.force_login(admin)
    
    print("Testing bulk uploads...")
    
    # Test 1: Examiner (with null fields)
    print("\n1. EXAMINER UPLOAD (with null fields)...")
    wb = Workbook()
    ws = wb.active
    headers = ['Name', 'Surname', 'Title', 'Qualification', 'Affiliation', 'Street_address',
               'Cell_phone', 'Email', 'Number_of_students_supervised', 'Current_affiliation',
               'Number_publications', 'International_assessor', 'Academic_experience']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    
    # Row with nulls
    ws['A2'] = 'TestExam'
    ws['B2'] = 'One'
    ws['C2'] = 'Prof'
    ws['D2'] = 'PhD'
    ws['E2'] = 'Univ'
    ws['F2'] = 'St'
    ws['G2'] = '111'
    ws['H2'] = 'exam1@test.com'
    ws['I2'] = None  # NULL - should default to 0
    ws['J2'] = 'U'
    ws['K2'] = None  # NULL - should default to 0
    ws['L2'] = None  # NULL - should default to False
    ws['M2'] = None  # NULL - should default to 0
    
    f = BytesIO()
    wb.save(f)
    f.seek(0)
    f.name = 'exam.xlsx'
    r = client.post('/admin/onboard/examiners', {'file': f})
    count = ExamminerProfile.objects.filter(email='exam1@test.com').count()
    print(f"   Examiners created: {count} - {'PASS' if count > 0 else 'FAIL'}")
    
    # Test 2: Supervisor
    print("\n2. SUPERVISOR UPLOAD...")
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(['Title', 'Names', 'Surname', 'Contact Details', 'email'], 1):
        ws.cell(1, col, h)
    ws['A2'] = 'Prof'
    ws['B2'] = 'TestSuper'
    ws['C2'] = 'One'
    ws['D2'] = '222'
    ws['E2'] = 'super1@test.com'
    
    f = BytesIO()
    wb.save(f)
    f.seek(0)
    f.name = 'super.xlsx'
    r = client.post('/admin/onboard/scholars', {'file': f})
    count = SupervisorProfile.objects.filter(email='super1@test.com').count()
    print(f"   Supervisors created: {count} - {'PASS' if count > 0 else 'FAIL'}")
    
    # Test 3: Student
    print("\n3. STUDENT UPLOAD...")
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary'], 1):
        ws.cell(1, col, h)
    ws['A2'] = 'Mr'
    ws['B2'] = 'TestStud'
    ws['C2'] = 'One'
    ws['D2'] = '333'
    ws['E2'] = 'STU001'
    ws['F2'] = 'stud1@test.com'
    ws['G2'] = 'alt@test.com'
    
    f = BytesIO()
    wb.save(f)
    f.seek(0)
    f.name = 'stud.xlsx'
    r = client.post('/admin/onboard/students', {'file': f, 'block_id': 'B1'})
    count = StudentProfile.objects.filter(student_no='STU001').count()
    print(f"   Students created: {count} - {'PASS' if count > 0 else 'FAIL'}")
    
    print("\nDone!")


if __name__ == '__main__':
    test_uploads()
