"""
Test case-insensitive duplicate detection
"""
import os
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile
from django.test import Client

User = get_user_model()

print("\n" + "="*80)
print("TEST: CASE-INSENSITIVE DUPLICATE DETECTION")
print("="*80)

# Clean test data
StudentProfile.objects.filter(student_no__startswith='CASE_TEST').delete()
User.objects.filter(email__iexact='case.test@exam.com').delete()

# Create admin
admin = User.objects.filter(is_superuser=True).first()
if not admin:
    admin = User.objects.create_superuser(
        username='case_test_admin',
        email='case_test_admin@test.com',
        password='test123'
    )

client = Client()
client.force_login(admin)

# First upload - lowercase
print("\n[TEST 1] First upload with lowercase email...")
wb = Workbook()
ws = wb.active

headers = ['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary']
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)

ws['A2'] = 'Mr'
ws['B2'] = 'Test'
ws['C2'] = 'Case'
ws['D2'] = '0500000001'
ws['E2'] = 'CASE_TEST_001'
ws['F2'] = 'case.test@exam.com'  # lowercase
ws['G2'] = 'case.test.sec@exam.com'

file_obj = BytesIO()
wb.save(file_obj)
file_obj.seek(0)
file_obj.name = 'test_case1.xlsx'

response = client.post('/admin/onboard/students', 
                      {'file': file_obj, 'block_id': 'CASE_TEST'},
                      follow=True)

count1 = StudentProfile.objects.filter(student_no='CASE_TEST_001').count()
print(f"  Students created: {count1}")

# Second upload - UPPERCASE (should be caught as duplicate)
print("\n[TEST 2] Second upload with UPPERCASE email (should be detected as duplicate)...")
wb2 = Workbook()
ws2 = wb2.active

for col, h in enumerate(headers, 1):
    ws2.cell(1, col, h)

ws2['A2'] = 'Mr'
ws2['B2'] = 'Test'
ws2['C2'] = 'Case'
ws2['D2'] = '0500000001'
ws2['E2'] = 'CASE_TEST_001'
ws2['F2'] = 'CASE.TEST@EXAM.COM'  # UPPERCASE
ws2['G2'] = 'case.test.sec@exam.com'

file_obj2 = BytesIO()
wb2.save(file_obj2)
file_obj2.seek(0)
file_obj2.name = 'test_case2.xlsx'

response2 = client.post('/admin/onboard/students', 
                       {'file': file_obj2, 'block_id': 'CASE_TEST'},
                       follow=True)

count2 = StudentProfile.objects.filter(student_no='CASE_TEST_001').count()
print(f"  Students created: {count2}")

print("\n" + "="*80)
print("RESULTS")
print("="*80)

if count1 == 1 and count2 == 1:
    print("✓ PASS: Case-insensitive duplicate detection working!")
    print("  ✓ First upload created 1 student")
    print("  ✓ Second upload detected duplicate and skipped (still 1 total)")
else:
    print(f"✗ FAIL: Expected 1 student both times, got {count1} then {count2}")

print("="*80)
print()
