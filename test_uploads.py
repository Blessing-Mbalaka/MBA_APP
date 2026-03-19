"""
Test script for bulk upload handlers
"""
import os
import sys
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import Client
from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile, ExamminerProfile, SupervisorProfile

AUser = get_user_model()


def create_admin_user():
    """Create or get admin user"""
    admin_user = AUser.objects.filter(username='testadmin').first()
    if not admin_user:
        admin_user = AUser.objects.create_user(
            username='testadmin',
            email='testadmin@test.com',
            password='testpass123',
            user_type=AUser.UserType.ADMIN
        )
    return admin_user


def create_examiner_excel():
    """Create excel file with examiners data"""
    wb = Workbook()
    ws = wb.active
    
    # Headers
    headers = ['Name', 'Surname', 'Title', 'Qualification', 'Affiliation', 'Street_address',
               'Cell_phone', 'Email', 'Number_of_students_supervised', 'Current_affiliation',
               'Number_publications', 'International_assessor', 'Academic_experience']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Row 1
    ws['A2'] = 'John'
    ws['B2'] = 'Smith'
    ws['C2'] = 'Prof'
    ws['D2'] = 'PhD'
    ws['E2'] = 'Univ'
    ws['F2'] = '123 St'
    ws['G2'] = '0721234567'
    ws['H2'] = 'e_john@test.com'
    ws['I2'] = 5
    ws['J2'] = 'Univ'
    ws['K2'] = 25
    ws['L2'] = True  # International_assessor - boolean
    ws['M2'] = 15
    
    # Row 2
    ws['A3'] = 'Jane'
    ws['B3'] = 'Doe'
    ws['C3'] = 'Dr'
    ws['D3'] = 'PhD'
    ws['E3'] = 'Inst'
    ws['F3'] = '456 Ave'
    ws['G3'] = '0789876543'
    ws['H3'] = 'e_jane@test.com'
    ws['I3'] = 3
    ws['J3'] = 'Univ'
    ws['K3'] = 18
    ws['L3'] = False  # International_assessor - boolean
    ws['M3'] = 12
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    file_obj.name = 'examiners.xlsx'
    return file_obj


def create_supervisor_excel():
    """Create excel file with supervisors"""
    wb = Workbook()
    ws = wb.active
    
    headers = ['Title', 'Names', 'Surname', 'Contact Details', 'email']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    ws['A2'] = 'Prof'
    ws['B2'] = 'Michael'
    ws['C2'] = 'Johnson'
    ws['D2'] = '0711111111'
    ws['E2'] = 'sv_michael@test.com'
    
    ws['A3'] = 'Dr'
    ws['B3'] = 'Sarah'
    ws['C3'] = 'Williams'
    ws['D3'] = '0722222222'
    ws['E3'] = 'sv_sarah@test.com'
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    file_obj.name = 'supervisors.xlsx'
    return file_obj


def create_student_excel():
    """Create excel file with students"""
    wb = Workbook()
    ws = wb.active
    
    headers = ['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    ws['A2'] = 'Mr'
    ws['B2'] = 'Brown'
    ws['C2'] = 'Robert'
    ws['D2'] = '0733333333'
    ws['E2'] = 'STU001'
    ws['F2'] = 's_robert@test.com'
    ws['G2'] = 's_robert_sec@test.com'
    
    ws['A3'] = 'Ms'
    ws['B3'] = 'Davis'
    ws['C3'] = 'Emma'
    ws['D3'] = '0744444444'
    ws['E3'] = 'STU002'
    ws['F3'] = 's_emma@test.com'
    ws['G3'] = 's_emma_sec@test.com'
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    file_obj.name = 'students.xlsx'
    return file_obj


def test_uploads():
    """Run all upload tests"""
    print("="*60)
    print("BULK UPLOAD TEST SUITE")
    print("="*60)
    
    # Create admin user
    admin = create_admin_user()
    client = Client()
    client.force_login(admin)
    
    # Test 1: Examiner upload
    print("\n[TEST] Examiner Upload...")
    excel = create_examiner_excel()
    resp = client.post('/admin/onboard/examiners', {'file': excel})
    e_count = AUser.objects.filter(email__startswith='e_').count()
    status = "PASS" if e_count == 2 else "FAIL"
    print(f"Created {e_count} examiners - {status}")
    
    # Test 2: Supervisor upload
    print("\n[TEST] Supervisor Upload...")
    excel = create_supervisor_excel()
    resp = client.post('/admin/onboard/scholars', {'file': excel})
    sv_count = AUser.objects.filter(email__startswith='sv_').count()
    status = "PASS" if sv_count == 2 else "FAIL"
    print(f"Created {sv_count} supervisors - {status}")
    
    # Test 3: Student upload
    print("\n[TEST] Student Upload...")
    excel = create_student_excel()
    resp = client.post('/admin/onboard/students', {'file': excel, 'block_id': 'TEST'})
    s_count = AUser.objects.filter(email__startswith='s_').count()
    status = "PASS" if s_count == 2 else "FAIL"
    print(f"Created {s_count} students - {status}")
    
    print("\n" + "="*60)
    total_pass = (1 if e_count == 2 else 0) + (1 if sv_count == 2 else 0) + (1 if s_count == 2 else 0)
    print(f"RESULT: {total_pass}/3 tests passed")
    print("="*60)
    
    return total_pass == 3


if __name__ == '__main__':
    success = test_uploads()
    sys.exit(0 if success else 1)
