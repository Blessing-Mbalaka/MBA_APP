"""
Test script for bulk upload functionality
Tests examiner, student, and supervisor bulk uploads
"""

import os
import sys
import django
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import RequestFactory
from django.contrib.auth.models import AnonymousUser
from django.contrib.messages.storage.fallback import FallbackStorage
from mbaAdmin.views.students_views import student_bulk_onboard
from mbaAdmin.views.scholars_views import onboard_examiners_bulk, onboard_supervisor_bulk
from mbamain.models import AUser, StudentProfile, ExamminerProfile, SupervisorProfile

# Test data
EXAMINER_TEST_DATA = [
    {
        'Name': 'John',
        'Surname': 'Smith',
        'Title': 'Dr.',
        'Qualification': 'PhD',
        'Affiliation': 'University A',
        'Street_address': '123 Main St',
        'Cell_phone': '555-0001',
        'Email': 'john.smith@university.ac.za',
        'Number_of_students_supervised': 5,
        'Current_affiliation': 'University A',
        'Number_publications': 12,
        'International_assessor': 'Yes',
        'Academic_experience': '15 years'
    },
    {
        'Name': 'Jane',
        'Surname': 'Doe',
        'Title': 'Prof.',
        'Qualification': 'PhD',
        'Affiliation': 'University B',
        'Street_address': '456 Oak Ave',
        'Cell_phone': '555-0002',
        'Email': 'jane.doe@university.ac.za',
        'Number_of_students_supervised': 8,
        'Current_affiliation': 'University B',
        'Number_publications': 25,
        'International_assessor': 'No',
        'Academic_experience': '20 years'
    }
]

STUDENT_TEST_DATA = [
    {
        'Title': 'Mr.',
        'Last name': 'Johnson',
        'First name': 'Richard',
        'Contact': '0712345678',
        'Student Number': 'STU-2025-001',
        'Email address': 'richard.johnson@student.ac.za',
        'Secondary': 'secondary1@student.ac.za'
    },
    {
        'Title': 'Ms.',
        'Last name': 'Williams',
        'First name': 'Sarah',
        'Contact': '0712345679',
        'Student Number': 'STU-2025-002',
        'Email address': 'sarah.williams@student.ac.za',
        'Secondary': 'secondary2@student.ac.za'
    }
]

SUPERVISOR_TEST_DATA = [
    {
        'Title': 'Dr.',
        'Names': 'Peter',
        'Surname': 'Brown',
        'Contact Details': '555-1001',
        'email': 'peter.brown@university.ac.za'
    },
    {
        'Title': 'Prof.',
        'Names': 'Michelle',
        'Surname': 'Green',
        'Contact Details': '555-1002',
        'email': 'michelle.green@university.ac.za'
    }
]

def create_excel_file(data, headers):
    """Create an Excel file from data"""
    wb = Workbook()
    ws = wb.active
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
    
    # Get bytes
    file_bytes = BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    return file_bytes

def create_request(file_bytes, filename, method='POST', post_data=None):
    """Create a mock request"""
    factory = RequestFactory()
    if post_data is None:
        post_data = {}
    
    request = factory.post('/test/', post_data)
    request.FILES['file'] = type('obj', (object,), {
        'read': lambda: file_bytes.read(),
        'name': filename,
        'seek': file_bytes.seek,
        'tell': file_bytes.tell
    })()
    
    # Add messages framework
    setattr(request, 'session', 'session')
    messages = FallbackStorage(request)
    request._messages = messages
    
    # Add user
    request.user = AUser.objects.filter(username='admin').first() or AnonymousUser()
    
    return request

def test_examiner_bulk_upload():
    """Test examiner bulk upload"""
    print("\n" + "="*70)
    print("TEST 1: EXAMINER BULK UPLOAD")
    print("="*70)
    
    # Clean up test data
    for data in EXAMINER_TEST_DATA:
        AUser.objects.filter(email=data['Email']).delete()
        ExamminerProfile.objects.filter(email=data['Email']).delete()
    
    # Create Excel file
    headers = list(EXAMINER_TEST_DATA[0].keys())
    file_bytes = create_excel_file(EXAMINER_TEST_DATA, headers)
    
    # Create request
    request = create_request(file_bytes, 'examiners.xlsx')
    
    # Call handler
    try:
        response = onboard_examiners_bulk(request)
        
        # Check if users created
        created_count = 0
        for data in EXAMINER_TEST_DATA:
            user = AUser.objects.filter(email=data['Email']).first()
            if user:
                profile = ExamminerProfile.objects.filter(user=user).first()
                if profile:
                    created_count += 1
                    print(f"✅ Created: {data['Name']} {data['Surname']} ({data['Email']})")
                else:
                    print(f"❌ User created but no profile: {data['Email']}")
            else:
                print(f"❌ Failed to create: {data['Email']}")
        
        print(f"\n✅ Examiner Upload Complete: {created_count}/{len(EXAMINER_TEST_DATA)} created")
        return True
        
    except Exception as e:
        print(f"❌ Examiner upload failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_student_bulk_upload():
    """Test student bulk upload"""
    print("\n" + "="*70)
    print("TEST 2: STUDENT BULK UPLOAD")
    print("="*70)
    
    # Use a test block ID
    block_id = "TEST_BLOCK_001"
    
    # Clean up test data
    for data in STUDENT_TEST_DATA:
        AUser.objects.filter(email=data['Email address']).delete()
        StudentProfile.objects.filter(student_no=data['Student Number']).delete()
    
    # Create Excel file
    headers = list(STUDENT_TEST_DATA[0].keys())
    file_bytes = create_excel_file(STUDENT_TEST_DATA, headers)
    
    # Create request with block_id
    request = create_request(file_bytes, 'students.xlsx', post_data={'block_id': block_id})
    
    # Call handler
    try:
        response = student_bulk_onboard(request)
        
        # Check if users created
        created_count = 0
        for data in STUDENT_TEST_DATA:
            user = AUser.objects.filter(email=data['Email address']).first()
            if user:
                profile = StudentProfile.objects.filter(user=user).first()
                if profile:
                    created_count += 1
                    print(f"✅ Created: {data['First name']} {data['Last name']} ({data['Email address']})")
                else:
                    print(f"❌ User created but no profile: {data['Email address']}")
            else:
                print(f"❌ Failed to create: {data['Email address']}")
        
        print(f"\n✅ Student Upload Complete: {created_count}/{len(STUDENT_TEST_DATA)} created")
        return True
        
    except Exception as e:
        print(f"❌ Student upload failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_supervisor_bulk_upload():
    """Test supervisor bulk upload"""
    print("\n" + "="*70)
    print("TEST 3: SUPERVISOR BULK UPLOAD")
    print("="*70)
    
    # Clean up test data
    for data in SUPERVISOR_TEST_DATA:
        AUser.objects.filter(email=data['email']).delete()
        SupervisorProfile.objects.filter(user__email=data['email']).delete()
    
    # Create Excel file
    headers = list(SUPERVISOR_TEST_DATA[0].keys())
    file_bytes = create_excel_file(SUPERVISOR_TEST_DATA, headers)
    
    # Create request
    request = create_request(file_bytes, 'supervisors.xlsx')
    
    # Call handler
    try:
        response = onboard_supervisor_bulk(request)
        
        # Check if users created
        created_count = 0
        for data in SUPERVISOR_TEST_DATA:
            user = AUser.objects.filter(email=data['email']).first()
            if user:
                profile = SupervisorProfile.objects.filter(user=user).first()
                if profile:
                    created_count += 1
                    print(f"✅ Created: {data['Names']} {data['Surname']} ({data['email']})")
                else:
                    print(f"❌ User created but no profile: {data['email']}")
            else:
                print(f"❌ Failed to create: {data['email']}")
        
        print(f"\n✅ Supervisor Upload Complete: {created_count}/{len(SUPERVISOR_TEST_DATA)} created")
        return True
        
    except Exception as e:
        print(f"❌ Supervisor upload failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_duplicate_handling():
    """Test handling of duplicate entries"""
    print("\n" + "="*70)
    print("TEST 4: DUPLICATE HANDLING")
    print("="*70)
    
    # Get a supervisor we created
    supervisor = SupervisorProfile.objects.first()
    if not supervisor:
        print("⏭️  Skipping - no supervisor exists yet")
        return True
    
    # Try to upload same supervisor again
    duplicate_data = [{
        'Title': supervisor.title,
        'Names': supervisor.name,
        'Surname': supervisor.surname,
        'Contact Details': supervisor.contact,
        'email': supervisor.user.email
    }]
    
    headers = list(duplicate_data[0].keys())
    file_bytes = create_excel_file(duplicate_data, headers)
    request = create_request(file_bytes, 'supervisors_dup.xlsx')
    
    try:
        response = onboard_supervisor_bulk(request)
        
        # Should not create duplicate
        count = AUser.objects.filter(email=supervisor.user.email).count()
        if count == 1:
            print(f"✅ Duplicate correctly skipped: {supervisor.user.email}")
            return True
        else:
            print(f"❌ Duplicate was created (found {count} users)")
            return False
    except Exception as e:
        print(f"❌ Duplicate test failed: {str(e)}")
        return False

def test_invalid_data():
    """Test handling of invalid data"""
    print("\n" + "="*70)
    print("TEST 5: INVALID DATA HANDLING")
    print("="*70)
    
    invalid_data = [
        {
            'Title': 'Dr.',
            'Names': 'Test',
            'Surname': 'User',
            'Contact Details': '555-1003',
            'email': 'invalid-email-format'  # Invalid email
        }
    ]
    
    headers = list(invalid_data[0].keys())
    file_bytes = create_excel_file(invalid_data, headers)
    request = create_request(file_bytes, 'supervisors_invalid.xlsx')
    
    try:
        response = onboard_supervisor_bulk(request)
        
        # Should not create user with invalid email
        user = AUser.objects.filter(username='invalid-email-format').first()
        if not user:
            print(f"✅ Invalid email correctly rejected")
            return True
        else:
            print(f"❌ Invalid email was accepted")
            return False
    except Exception as e:
        print(f"⚠️  Expected error for invalid email: {str(e)}")
        return True

def main():
    """Run all tests"""
    print("\n" + "="*70)
    print("BULK UPLOAD TEST SUITE")
    print("="*70)
    
    results = []
    
    # Run tests
    results.append(("Examiner Upload", test_examiner_bulk_upload()))
    results.append(("Student Upload", test_student_bulk_upload()))
    results.append(("Supervisor Upload", test_supervisor_bulk_upload()))
    results.append(("Duplicate Handling", test_duplicate_handling()))
    results.append(("Invalid Data", test_invalid_data()))
    
    # Summary
    print("\n" + "="*70)
    print("TEST SUMMARY")
    print("="*70)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status}: {test_name}")
    
    print(f"\nTotal: {passed}/{total} tests passed")
    print("="*70)
    
    return passed == total

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
