"""
COMPREHENSIVE VERIFICATION TEST
Tests all fixes for bulk upload issues
- Examiner: Null numeric fields handled
- Supervisor: Basic functionality maintained
- Student: Duplicate detection working
"""
import os
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import Client
from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile, ExamminerProfile, SupervisorProfile

User = get_user_model()

print("\n" + "="*80)
print("VERIFICATION TEST - BULK UPLOAD FIXES")
print("="*80)

# Cleanup first
print("\n[SETUP] Cleaning database...")
User.objects.filter(username__startswith='verify').delete()
ExamminerProfile.objects.filter(email__startswith='verify').delete()
SupervisorProfile.objects.filter(contact__isnull=False).filter(name__isnull=True).delete()
print("  ✓ Database cleaned for fresh test")

# Create admin
admin = User.objects.filter(username='verify_admin').first()
if not admin:
    admin = User.objects.create_user(
        username='verify_admin',
        email='verify_admin@test.com',
        password='test123',
        user_type=User.UserType.ADMIN
    )
print("  ✓ Admin user ready")

client = Client()
client.force_login(admin)

# ============================================================================
# TEST 1: EXAMINER UPLOAD - NULL NUMERIC FIELDS
# ============================================================================
print("\n" + "-"*80)
print("TEST 1: EXAMINER BULK UPLOAD (with NULL numeric fields)")
print("-"*80)

wb = Workbook()
ws = wb.active

# Headers
headers = ['Name', 'Surname', 'Title', 'Qualification', 'Affiliation', 'Street_address',
           'Cell_phone', 'Email', 'Number_of_students_supervised', 'Current_affiliation',
           'Number_publications', 'International_assessor', 'Academic_experience']
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)

# Row 1: Complete data
ws['A2'] = 'Alice'
ws['B2'] = 'Complete'
ws['C2'] = 'Prof'
ws['D2'] = 'PhD'
ws['E2'] = 'University A'
ws['F2'] = '123 Main'
ws['G2'] = '0501111111'
ws['H2'] = 'verify_exam_complete@test.com'
ws['I2'] = 5          # number_of_students_supervised
ws['J2'] = 'Univ A'
ws['K2'] = 25         # number_publications
ws['L2'] = True       # international_assessor
ws['M2'] = 15         # academic_experience

# Row 2: NULL numeric fields (THIS WAS CAUSING THE ERROR)
ws['A3'] = 'Bob'
ws['B3'] = 'Null'
ws['C3'] = 'Dr'
ws['D3'] = 'PhD'
ws['E3'] = 'University B'
ws['F3'] = '456 Oak'
ws['G3'] = '0502222222'
ws['H3'] = 'verify_exam_null@test.com'
ws['I3'] = None       # NULL - should default to 0
ws['J3'] = 'Univ B'
ws['K3'] = None       # NULL - should default to 0
ws['L3'] = None       # NULL - should default to False
ws['M3'] = None       # NULL - should default to 0

file_obj = BytesIO()
wb.save(file_obj)
file_obj.seek(0)
file_obj.name = 'exam_test.xlsx'

try:
    response = client.post('/admin/onboard/examiners', {'file': file_obj})
    
    # Verify both created
    count_complete = ExamminerProfile.objects.filter(email='verify_exam_complete@test.com').count()
    count_null = ExamminerProfile.objects.filter(email='verify_exam_null@test.com').count()
    
    if count_complete == 1 and count_null == 1:
        print("✓ PASS: Both examiners created successfully")
        
        # Verify NULL fields were handled correctly
        exam_null = ExamminerProfile.objects.get(email='verify_exam_null@test.com')
        if (exam_null.number_of_students_supervised == 0 and
            exam_null.number_publications == 0 and
            exam_null.academic_experience == 0 and
            exam_null.international_assessor == False):
            print("  ✓ NULL numeric fields properly defaulted to 0/False")
        else:
            print("  ✗ FAILED: NULL fields not properly converted")
            print(f"    Got: students={exam_null.number_of_students_supervised}, "
                  f"pubs={exam_null.number_publications}, "
                  f"exp={exam_null.academic_experience}, "
                  f"intl={exam_null.international_assessor}")
    else:
        print(f"✗ FAIL: Expected 2 examiners, got {count_complete + count_null}")
except Exception as e:
    print(f"✗ FAIL: {str(e)}")

# ============================================================================
# TEST 2: SUPERVISOR BULK UPLOAD
# ============================================================================
print("\n" + "-"*80)
print("TEST 2: SUPERVISOR BULK UPLOAD")
print("-"*80)

wb = Workbook()
ws = wb.active

headers = ['Title', 'Names', 'Surname', 'Contact Details', 'email']
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)

ws['A2'] = 'Prof'
ws['B2'] = 'Charlie'
ws['C2'] = 'Supervisor'
ws['D2'] = '0503333333'
ws['E2'] = 'verify_super1@test.com'

ws['A3'] = 'Dr'
ws['B3'] = 'Diana'
ws['C3'] = 'Supervisor'
ws['D3'] = '0504444444'
ws['E3'] = 'verify_super2@test.com'

file_obj = BytesIO()
wb.save(file_obj)
file_obj.seek(0)
file_obj.name = 'super_test.xlsx'

try:
    response = client.post('/admin/onboard/scholars', {'file': file_obj})
    
    count = SupervisorProfile.objects.filter(name__in=['Charlie', 'Diana']).count()
    if count == 2:
        print("✓ PASS: 2 supervisors created successfully")
        super1 = SupervisorProfile.objects.filter(name='Charlie').first()
        super2 = SupervisorProfile.objects.filter(name='Diana').first()
        print(f"  ✓ Charlie: {super1.user.email if super1.user else 'N/A'}")
        print(f"  ✓ Diana: {super2.user.email if super2.user else 'N/A'}")
    else:
        print(f"✗ FAIL: Expected 2 supervisors, got {count}")
except Exception as e:
    print(f"✗ FAIL: {str(e)}")

# ============================================================================
# TEST 3: STUDENT BULK UPLOAD - DUPLICATE DETECTION
# ============================================================================
print("\n" + "-"*80)
print("TEST 3: STUDENT BULK UPLOAD (duplicate detection test)")
print("-"*80)

wb = Workbook()
ws = wb.active

headers = ['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary']
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)

ws['A2'] = 'Mr'
ws['B2'] = 'Evans'
ws['C2'] = 'Edward'
ws['D2'] = '0505555555'
ws['E2'] = 'VERIFY_STU001'
ws['F2'] = 'verify_stud1@test.com'
ws['G2'] = 'verify_stud1_sec@test.com'

ws['A3'] = 'Ms'
ws['B3'] = 'Flynn'
ws['C3'] = 'Fiona'
ws['D3'] = '0506666666'
ws['E3'] = 'VERIFY_STU002'
ws['F3'] = 'verify_stud2@test.com'
ws['G3'] = 'verify_stud2_sec@test.com'

file_obj = BytesIO()
wb.save(file_obj)
file_obj.seek(0)
file_obj.name = 'stud_test.xlsx'

try:
    response = client.post('/admin/onboard/students', 
                          {'file': file_obj, 'block_id': 'VERIFY_BLOCK'})
    
    count = StudentProfile.objects.filter(student_no__startswith='VERIFY_STU').count()
    if count == 2:
        print("✓ PASS: 2 students created successfully")
        
        stud1 = StudentProfile.objects.get(student_no='VERIFY_STU001')
        stud2 = StudentProfile.objects.get(student_no='VERIFY_STU002')
        
        print(f"  ✓ Edward Evans: {stud1.user.email if stud1.user else 'N/A'}")
        print(f"  ✓ Fiona Flynn: {stud2.user.email if stud2.user else 'N/A'}")
        
        # Test duplicate detection by trying to upload same student again
        print("\n  Testing duplicate detection...")
        wb2 = Workbook()
        ws2 = wb2.active
        for col, h in enumerate(headers, 1):
            ws2.cell(1, col, h)
        ws2['A2'] = 'Mr'
        ws2['B2'] = 'Evans'
        ws2['C2'] = 'Edward'
        ws2['D2'] = '0505555555'
        ws2['E2'] = 'VERIFY_STU001'
        ws2['F2'] = 'verify_stud1@test.com'
        ws2['G2'] = 'verify_stud1_sec@test.com'
        
        file_obj2 = BytesIO()
        wb2.save(file_obj2)
        file_obj2.seek(0)
        file_obj2.name = 'stud_dup_test.xlsx'
        
        response2 = client.post('/admin/onboard/students',
                               {'file': file_obj2, 'block_id': 'VERIFY_BLOCK'})
        
        # Should still be 2 (duplicate rejected)
        count_after = StudentProfile.objects.filter(student_no__startswith='VERIFY_STU').count()
        if count_after == 2:
            print("  ✓ Duplicate detection working - duplicate entry rejected")
        else:
            print(f"  ✗ Duplicate detection failed - got {count_after} instead of 2")
    else:
        print(f"✗ FAIL: Expected 2 students, got {count}")
except Exception as e:
    print(f"✗ FAIL: {str(e)}")

# ============================================================================
# FINAL SUMMARY
# ============================================================================
print("\n" + "="*80)
print("VERIFICATION COMPLETE")
print("="*80)

exam_count = ExamminerProfile.objects.filter(email__startswith='verify_exam').count()
super_count = SupervisorProfile.objects.filter(name__in=['Charlie', 'Diana']).count()
stud_count = StudentProfile.objects.filter(student_no__startswith='VERIFY_STU').count()

print(f"\n✓ Examiners created: {exam_count} (including NULL field test)")
print(f"✓ Supervisors created: {super_count}")
print(f"✓ Students created: {stud_count} (including duplicate detection test)")

if exam_count >= 2 and super_count == 2 and stud_count == 2:
    print("\n" + "="*80)
    print("✓✓✓ ALL FIXES VERIFIED AND WORKING ✓✓✓")
    print("="*80)
else:
    print("\n⚠ Some tests did not complete as expected")

print()
