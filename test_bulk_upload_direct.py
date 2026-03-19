"""
Direct test script for bulk upload functionality
Tests without mocking the request object
"""

import os
import sys
import django
from pathlib import Path
from openpyxl import Workbook
from io import BytesIO

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import Client
from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile, ExamminerProfile, SupervisorProfile

User = get_user_model()

# Test data
EXAMINER_TEST_DATA = [
    ('John', 'Smith', 'Dr.', 'PhD', 'University A', '123 Main St', '555-0001', 
     'e_john.smith@test.ac.za', 5, 'University A', 12, True, 15),
    ('Jane', 'Doe', 'Prof.', 'PhD', 'University B', '456 Oak Ave', '555-0002',
     'e_jane.doe@test.ac.za', 8, 'University B', 25, False, 20),
]

STUDENT_TEST_DATA = [
    ('Mr.', 'Johnson', 'Richard', '0712345678', 'STU-2025-001',
     's_richard.johnson@test.ac.za', 'secondary1@test.ac.za'),
    ('Ms.', 'Williams', 'Sarah', '0712345679', 'STU-2025-002',
     's_sarah.williams@test.ac.za', 'secondary2@test.ac.za'),
]

SUPERVISOR_TEST_DATA = [
    ('Dr.', 'Peter', 'Brown', '555-1001', 'sv_peter.brown@test.ac.za'),
    ('Prof.', 'Michelle', 'Green', '555-1002', 'sv_michelle.green@test.ac.za'),
]

def create_excel_file(data_list, headers):
    """Create an Excel file from data"""
    wb = Workbook()
    ws = wb.active
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_num, row_data in enumerate(data_list, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Get bytes
    file_bytes = BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    return file_bytes

def login_as_admin(client):
    """Login as admin user"""
    # Create admin if doesn't exist
    admin_user = User.objects.filter(username='admin').first()
    if not admin_user:
        print("Creating admin user...")
        admin_user = User.objects.create_user(
            username='admin',
            email='admin@test.ac.za',
            password='admin123'
        )
        admin_user.set_user_type(User.UserType.ADMIN)
        admin_user.save()
    
    # Login
    login_result = client.login(username='admin', password='admin123')
    if not login_result:
        print("⚠️  Could not login as admin - trying direct auth")
    return client

def test_examiner_upload():
    """Test examiner bulk upload"""
    print("\n" + "="*70)
    print("TEST 1: EXAMINER BULK UPLOAD")
    print("="*70)
    
    # Clean up
    for data in EXAMINER_TEST_DATA:
        User.objects.filter(email=data[7]).delete()
    
    client = login_as_admin(Client())
    
    # Create Excel file
    headers = ['Name', 'Surname', 'Title', 'Qualification', 'Affiliation', 
               'Street_address', 'Cell_phone', 'Email', 'Number_of_students_supervised',
               'Current_affiliation', 'Number_publications', 'International_assessor', 
               'Academic_experience']
    file_bytes = create_excel_file(EXAMINER_TEST_DATA, headers)
    file_bytes.name = 'examiners.xlsx'
    
    # Upload
    try:
        response = client.post('/admin/onboard/examiners', data={'file': file_bytes})
        
        # Check results
        created = 0
        for data in EXAMINER_TEST_DATA:
            user = User.objects.filter(email=data[7]).first()
            if user:
                profile = ExamminerProfile.objects.filter(user=user).first()
                if profile:
                    created += 1
                    print(f"✅ Created: {data[0]} {data[1]} ({data[7]})")
                else:
                    print(f"❌ User created but no profile: {data[7]}")
            else:
                print(f"❌ Not created: {data[7]}")
        
        print(f"Result: {created}/{len(EXAMINER_TEST_DATA)} examiners created")
        return created > 0
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_student_upload():
    """Test student bulk upload"""
    print("\n" + "="*70)
    print("TEST 2: STUDENT BULK UPLOAD")
    print("="*70)
    
    # Clean up
    for data in STUDENT_TEST_DATA:
        User.objects.filter(email=data[5]).delete()
    
    client = login_as_admin(Client())
    
    # Create Excel file
    headers = ['Title', 'Last name', 'First name', 'Contact', 'Student Number',
               'Email address', 'Secondary']
    file_bytes = create_excel_file(STUDENT_TEST_DATA, headers)
    file_bytes.name = 'students.xlsx'
    
    # Upload with block_id
    try:
        response = client.post('/admin/onboard/students', data={
            'file': file_bytes,
            'block_id': 'TEST_BLOCK'
        })
        
        # Check results
        created = 0
        for data in STUDENT_TEST_DATA:
            user = User.objects.filter(email=data[5]).first()
            if user:
                profile = StudentProfile.objects.filter(user=user).first()
                if profile:
                    created += 1
                    print(f"✅ Created: {data[2]} {data[1]} ({data[5]})")
                else:
                    print(f"❌ User created but no profile: {data[5]}")
            else:
                print(f"❌ Not created: {data[5]}")
        
        print(f"Result: {created}/{len(STUDENT_TEST_DATA)} students created")
        return created > 0
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_supervisor_upload():
    """Test supervisor bulk upload"""
    print("\n" + "="*70)
    print("TEST 3: SUPERVISOR BULK UPLOAD")
    print("="*70)
    
    # Clean up
    for data in SUPERVISOR_TEST_DATA:
        User.objects.filter(email=data[4]).delete()
    
    client = login_as_admin(Client())
    
    # Create Excel file
    headers = ['Title', 'Names', 'Surname', 'Contact Details', 'email']
    file_bytes = create_excel_file(SUPERVISOR_TEST_DATA, headers)
    file_bytes.name = 'supervisors.xlsx'
    
    # Upload
    try:
        response = client.post('/admin/onboard/scholars', data={'file': file_bytes})
        
        # Check results
        created = 0
        for data in SUPERVISOR_TEST_DATA:
            user = User.objects.filter(email=data[4]).first()
            if user:
                profile = SupervisorProfile.objects.filter(user=user).first()
                if profile:
                    created += 1
                    print(f"✅ Created: {data[1]} {data[2]} ({data[4]})")
                else:
                    print(f"❌ User created but no profile: {data[4]}")
            else:
                print(f"❌ Not created: {data[4]}")
        
        print(f"Result: {created}/{len(SUPERVISOR_TEST_DATA)} supervisors created")
        return created > 0
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Run all tests"""
    print("\n" + "="*70)
    print("BULK UPLOAD TEST SUITE - DIRECT")
    print("="*70)
    
    results = []
    
    try:
        results.append(("Examiner Upload", test_examiner_upload()))
        results.append(("Student Upload", test_student_upload()))
        results.append(("Supervisor Upload", test_supervisor_upload()))
    except Exception as e:
        print(f"\n❌ Test suite error: {str(e)}")
        import traceback
        traceback.print_exc()
    
    # Summary
    print("\n" + "="*70)
    print("TEST SUMMARY")
    print("="*70)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status}: {test_name}")
    
    print(f"\nTotal: {passed}/{total} tests passed")
    print("="*70)
    
    return passed == total

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
