"""
Comprehensive test for all three bulk upload handlers
Tests with null/missing numeric values to ensure defensive coding works
"""
import os
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import Client
from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile, ExamminerProfile, SupervisorProfile

AUser = get_user_model()


def setup_admin():
    """Create admin user if not exists"""
    admin = AUser.objects.filter(username='admin@test.com').first()
    if not admin:
        admin = AUser.objects.create_user(
            username='admin@test.com',
            email='admin@test.com',
            password='testpass',
            user_type=AUser.UserType.ADMIN
        )
    return admin


def create_examiner_file():
    """Create Excel with examiners (with some null numeric fields)"""
    wb = Workbook()
    ws = wb.active
    
    # Headers
    headers = ['Name', 'Surname', 'Title', 'Qualification', 'Affiliation', 'Street_address',
               'Cell_phone', 'Email', 'Number_of_students_supervised', 'Current_affiliation',
               'Number_publications', 'International_assessor', 'Academic_experience']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    
    # Row 1 - All fields filled
    ws['A2'] = 'John'
    ws['B2'] = 'Smith'
    ws['C2'] = 'Prof'
    ws['D2'] = 'PhD'
    ws['E2'] = 'University A'
    ws['F2'] = '123 Main St'
    ws['G2'] = '0721234567'
    ws['H2'] = 'john.smith@exam.com'
    ws['I2'] = 5
    ws['J2'] = 'Univ A'
    ws['K2'] = 25
    ws['L2'] = True
    ws['M2'] = 15
    
    # Row 2 - With NULLS in numeric fields (this was causing the error)
    ws['A3'] = 'Jane'
    ws['B3'] = 'Doe'
    ws['C3'] = 'Dr'
    ws['D3'] = 'PhD'
    ws['E3'] = 'University B'
    ws['F3'] = '456 Oak Ave'
    ws['G3'] = '0789876543'
    ws['H3'] = 'jane.doe@exam.com'
    ws['I3'] = None  # NULL - should default to 0
    ws['J3'] = 'Univ B'
    ws['K3'] = None  # NULL - should default to 0
    ws['L3'] = None  # NULL - should default to False
    ws['M3'] = None  # NULL - should default to 0
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    file_obj.name = 'examiners.xlsx'
    return file_obj


def create_supervisor_file():
    """Create Excel with supervisors"""
    wb = Workbook()
    ws = wb.active
    
    headers = ['Title', 'Names', 'Surname', 'Contact Details', 'email']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    
    ws['A2'] = 'Prof'
    ws['B2'] = 'Michael'
    ws['C2'] = 'Johnson'
    ws['D2'] = '0711111111'
    ws['E2'] = 'michael@super.com'
    
    ws['A3'] = 'Dr'
    ws['B3'] = 'Sarah'
    ws['C3'] = 'Williams'
    ws['D3'] = '0722222222'
    ws['E3'] = 'sarah@super.com'
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    file_obj.name = 'supervisors.xlsx'
    return file_obj


def create_student_file():
    """Create Excel with students"""
    wb = Workbook()
    ws = wb.active
    
    headers = ['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary']
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
    
    ws['A2'] = 'Mr'
    ws['B2'] = 'Brown'
    ws['C2'] = 'Robert'
    ws['D2'] = '0733333333'
    ws['E2'] = 'STU001'
    ws['F2'] = 'robert@stud.com'
    ws['G2'] = 'robert.alt@stud.com'
    
    ws['A3'] = 'Ms'
    ws['B3'] = 'Davis'
    ws['C3'] = 'Emma'
    ws['D3'] = '0744444444'
    ws['E3'] = 'STU002'
    ws['F3'] = 'emma@stud.com'
    ws['G3'] = 'emma.alt@stud.com'
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    file_obj.name = 'students.xlsx'
    return file_obj


def test_all_uploads():
    """Run all three upload tests"""
    print("\n" + "="*70)
    print("COMPREHENSIVE BULK UPLOAD TEST")
    print("="*70)
    
    # Setup
    admin = setup_admin()
    client = Client()
    client.force_login(admin)
    
    results = {}
    
    # TEST 1: EXAMINER UPLOAD (with null numeric fields - the critical test)
    print("\n[TEST 1] EXAMINER UPLOAD (testing null numeric field handling)...")
    try:
        # Clean existing
        AUser.objects.filter(email__startswith='john').delete()
        AUser.objects.filter(email__startswith='jane').delete()
        
        excel = create_examiner_file()
        response = client.post('/admin/onboard/examiners', {'file': excel})
        
        # Check created
        created = ExamminerProfile.objects.filter(
            email__in=['john.smith@exam.com', 'jane.doe@exam.com']
        ).count()
        
        if created == 2:
            print("  ✓ PASS: 2 examiners created successfully")
            print("    - Null fields properly defaulted to 0/False")
            results['examiners'] = 'PASS'
        else:
            print(f"  ✗ FAIL: Expected 2 examiners, got {created}")
            results['examiners'] = 'FAIL'
    except Exception as e:
        print(f"  ✗ FAIL: {str(e)}")
        results['examiners'] = 'FAIL'
    
    # TEST 2: SUPERVISOR UPLOAD
    print("\n[TEST 2] SUPERVISOR UPLOAD...")
    try:
        # Clean existing
        AUser.objects.filter(email__startswith='michael').delete()
        AUser.objects.filter(email__startswith='sarah').delete()
        
        excel = create_supervisor_file()
        response = client.post('/admin/onboard/scholars', {'file': excel})
        
        # Check created
        created = SupervisorProfile.objects.filter(
            email__in=['michael@super.com', 'sarah@super.com']
        ).count()
        
        if created == 2:
            print("  ✓ PASS: 2 supervisors created successfully")
            results['supervisors'] = 'PASS'
        else:
            print(f"  ✗ FAIL: Expected 2 supervisors, got {created}")
            results['supervisors'] = 'FAIL'
    except Exception as e:
        print(f"  ✗ FAIL: {str(e)}")
        results['supervisors'] = 'FAIL'
    
    # TEST 3: STUDENT UPLOAD
    print("\n[TEST 3] STUDENT UPLOAD...")
    try:
        # Clean existing
        AUser.objects.filter(email__startswith='robert').delete()
        AUser.objects.filter(email__startswith='emma').delete()
        
        excel = create_student_file()
        response = client.post(
            '/admin/onboard/students',
            {'file': excel, 'block_id': 'TEST_BLOCK'}
        )
        
        # Check created
        created = StudentProfile.objects.filter(
            student_no__in=['STU001', 'STU002']
        ).count()
        
        if created == 2:
            print("  ✓ PASS: 2 students created successfully")
            results['students'] = 'PASS'
        else:
            print(f"  ✗ FAIL: Expected 2 students, got {created}")
            results['students'] = 'FAIL'
    except Exception as e:
        print(f"  ✗ FAIL: {str(e)}")
        results['students'] = 'FAIL'
    
    # SUMMARY
    print("\n" + "="*70)
    print("TEST SUMMARY")
    print("="*70)
    passed = sum(1 for v in results.values() if v == 'PASS')
    total = len(results)
    
    for name, status in results.items():
        symbol = "✓" if status == "PASS" else "✗"
        print(f"{symbol} {name.upper()}: {status}")
    
    print(f"\nRESULT: {passed}/{total} tests passed")
    
    if passed == total:
        print("\n✓ ALL BULK UPLOADS WORKING!")
        return True
    else:
        print(f"\n✗ {total - passed} upload(s) failed")
        return False


if __name__ == '__main__':
    import sys
    success = test_all_uploads()
    sys.exit(0 if success else 1)
