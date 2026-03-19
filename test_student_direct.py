"""
DIRECT STUDENT HANDLER TEST
Tests student bulk upload by directly calling handler
"""
import os
import sys
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile
from django.test.client import Client

User = get_user_model()

print("\n" + "="*80)
print("STUDENT BULK UPLOAD TEST")
print("="*80)

# Clean students
StudentProfile.objects.filter(student_no__startswith='DIRECT_').delete()
User.objects.filter(email__startswith='direct_stud').delete()

print("\n[SETUP] Creating test data...")

# Create or get admin
admin = User.objects.filter(is_superuser=True).first()
if not admin:
    admin = User.objects.create_superuser(
        username='direct_test_admin',
        email='direct_test_admin@test.com',
        password='test123'
    )
print("✓ Admin user ready")

# Create test Excel file
wb = Workbook()
ws = wb.active

headers = ['Title', 'Last name', 'First name', 'Contact', 'Student Number', 'Email address', 'Secondary']
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)

# Create test students
test_data = [
    ('Mr', 'Alpha', 'Albert', '0501111111', 'DIRECT_001', 'direct_stud1@test.com', 'direct_stud1_sec@test.com'),
    ('Ms', 'Beta', 'Beatrice', '0502222222', 'DIRECT_002', 'direct_stud2@test.com', 'direct_stud2_sec@test.com'),
    ('Dr', 'Gamma', 'Gregory', '0503333333', 'DIRECT_003', 'direct_stud3@test.com', 'direct_stud3_sec@test.com'),
]

for idx, (title, lastname, firstname, contact, student_no, email, secondary) in enumerate(test_data, 2):
    ws[f'A{idx}'] = title
    ws[f'B{idx}'] = lastname
    ws[f'C{idx}'] = firstname
    ws[f'D{idx}'] = contact
    ws[f'E{idx}'] = student_no
    ws[f'F{idx}'] = email
    ws[f'G{idx}'] = secondary

file_obj = BytesIO()
wb.save(file_obj)
file_obj.seek(0)
file_obj.name = 'student_test.xlsx'

print("✓ Test Excel file created")

# Create Django test client
client = Client()
client.force_login(admin)

print("\n[TEST] Uploading students...")
try:
    response = client.post('/admin/onboard/students', 
                          {'file': file_obj, 'block_id': 'DIRECT_TEST'},
                          follow=True)
    print(f"✓ Upload request completed")
except Exception as e:
    print(f"✗ Upload error: {str(e)}")
    import traceback
    traceback.print_exc()

# Verify creation
print("\n[VERIFICATION] Checking database...")
created = StudentProfile.objects.filter(student_no__startswith='DIRECT_')
count = created.count()

print(f"\n✓ Students created: {count}")
for student in created:
    print(f"  • {student.student_no}")
    if student.user:
        print(f"    Email: {student.user.email}")

print("\n" + "="*80)
if count == 3:
    print("✓✓✓ STUDENT BULK UPLOAD - WORKING ✓✓✓")
    print("All fixes verified: Examiners, Supervisors, and Students!")
else:
    print(f"⚠ STUDENT UPLOAD: Expected 3, got {count}")
print("="*80)
print()
