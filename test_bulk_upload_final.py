"""
Comprehensive test suite for bulk upload handlers (ASCII version for Windows)
"""
import os
import sys
import django
from io import BytesIO
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mysite.settings')
django.setup()

from django.test import Client
from django.contrib.auth import get_user_model
from mbamain.models import StudentProfile, ExamminerProfile, SupervisorProfile

AUser = get_user_model()


def create_test_excel_examiner():
    """Create test examiner Excel file"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Surname'
    ws['C1'] = 'Title'
    ws['D1'] = 'Qualification'
    ws['E1'] = 'Affiliation'
    ws['F1'] = 'Street_address'
    ws['G1'] = 'Cell_phone'
    ws['H1'] = 'Email'
    ws['I1'] = 'Number_of_students_supervised'
    ws['J1'] = 'Current_affiliation'
    ws['K1'] = 'Number_publications'
    ws['L1'] = 'International_assessor'
    ws['M1'] = 'Academic_experience'
    
    ws['A2'] = 'John'
    ws['B2'] = 'Smith'
    ws['C2'] = 'Prof'
    ws['D2'] = 'PhD'
    ws['E2'] = 'University'
    ws['F2'] = '123 Main St'
    ws['G2'] = '0721234567'
    ws['H2'] = 'e_john_smith@test.com'
    ws['I2'] = 5
    ws['J2'] = 'Current Uni'
    ws['K2'] = 25
    ws['L2'] = 'Yes'
    ws['M2'] = 15
    
    ws['A3'] = 'Jane'
    ws['B3'] = 'Doe'
    ws['C3'] = 'Dr'
    ws['D3'] = 'PhD'
    ws['E3'] = 'Institute'
    ws['F3'] = '456 Second Ave'
    ws['G3'] = '0789876543'
    ws['H3'] = 'e_jane_doe@test.com'
    ws['I3'] = 3
    ws['J3'] = 'Other Uni'
    ws['K3'] = 18
    ws['L3'] = 'No'
    ws['M3'] = 12
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    return file_obj


def create_test_excel_supervisor():
    """Create test supervisor Excel file"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Title'
    ws['B1'] = 'Names'
    ws['C1'] = 'Surname'
    ws['D1'] = 'Contact Details'
    ws['E1'] = 'email'
    
    ws['A2'] = 'Prof'
    ws['B2'] = 'Michael'
    ws['C2'] = 'Johnson'
    ws['D2'] = '0711111111'
    ws['E2'] = 'sv_michael_johnson@test.com'
    
    ws['A3'] = 'Dr'
    ws['B3'] = 'Sarah'
    ws['C3'] = 'Williams'
    ws['D3'] = '0722222222'
    ws['E3'] = 'sv_sarah_williams@test.com'
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    return file_obj


def create_test_excel_student():
    """Create test student Excel file"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Title'
    ws['B1'] = 'Last name'
    ws['C1'] = 'First name'
    ws['D1'] = 'Contact'
    ws['E1'] = 'Student Number'
    ws['F1'] = 'Email address'
    ws['G1'] = 'Secondary'
    
    ws['A2'] = 'Mr'
    ws['B2'] = 'Brown'
    ws['C2'] = 'Robert'
    ws['D2'] = '0733333333'
    ws['E2'] = 'STU001'
    ws['F2'] = 's_robert_brown@test.com'
    ws['G2'] = 's_robert_brown_sec@test.com'
    
    ws['A3'] = 'Ms'
    ws['B3'] = 'Davis'
    ws['C3'] = 'Emma'
    ws['D3'] = '0744444444'
    ws['E3'] = 'STU002'
    ws['F3'] = 's_emma_davis@test.com'
    ws['G3'] = 's_emma_davis_sec@test.com'
    
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)
    return file_obj


def test_examiner_upload():
    """Test examiner bulk upload"""
    print("\n[TEST] Starting Examiner Upload Test...")
    try:
        # Create or get admin user
        admin_user, created = AUser.objects.get_or_create(
            username='admin_test',
            defaults={
                'email': 'admin_test@test.com',
                'user_type': AUser.UserType.ADMIN
            }
        )
        if created:
            admin_user.set_password('test_password')
            admin_user.save()
        
        excel_file = create_test_excel_examiner()
        excel_file.name = 'test_examiners.xlsx'
        
        client = Client()
        client.login(username='admin_test', password='test_password')
        
        response = client.post(
            '/admin/onboard/examiners',
            {'file': excel_file},
            follow=True
        )
        
        # Check if users were created
        examiner_users = AUser.objects.filter(email__startswith='e_')
        print(f"[EXAMINER] Created {examiner_users.count()} examiner users")
        
        for user in examiner_users:
            profile = ExamminerProfile.objects.filter(user=user)
            status = "OK" if profile.exists() else "MISSING PROFILE"
            print(f"  - {user.email}: {status}")
        
        if examiner_users.count() == 2:
            print("[PASS] Examiner upload successful")
            return True
        else:
            print("[FAIL] Expected 2 examiners, got " + str(examiner_users.count()))
            return False
            
    except Exception as e:
        print(f"[ERROR] Examiner test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def test_supervisor_upload():
    """Test supervisor bulk upload"""
    print("\n[TEST] Starting Supervisor Upload Test...")
    try:
        # Create or get admin user
        admin_user, created = AUser.objects.get_or_create(
            username='admin_test',
            defaults={
                'email': 'admin_test@test.com',
                'user_type': AUser.UserType.ADMIN
            }
        )
        if created:
            admin_user.set_password('test_password')
            admin_user.save()
        
        excel_file = create_test_excel_supervisor()
        excel_file.name = 'test_supervisors.xlsx'
        
        client = Client()
        client.login(username='admin_test', password='test_password')
        
        response = client.post(
            '/admin/onboard/scholars',
            {'file': excel_file},
            follow=True
        )
        
        # Check if users were created
        supervisor_users = AUser.objects.filter(email__startswith='sv_')
        print(f"[SUPERVISOR] Created {supervisor_users.count()} supervisor users")
        
        for user in supervisor_users:
            profile = SupervisorProfile.objects.filter(user=user)
            status = "OK" if profile.exists() else "MISSING PROFILE"
            print(f"  - {user.email}: {status}")
        
        if supervisor_users.count() == 2:
            print("[PASS] Supervisor upload successful")
            return True
        else:
            print("[FAIL] Expected 2 supervisors, got " + str(supervisor_users.count()))
            return False
            
    except Exception as e:
        print(f"[ERROR] Supervisor test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def test_student_upload():
    """Test student bulk upload"""
    print("\n[TEST] Starting Student Upload Test...")
    try:
        # Create or get admin user
        admin_user, created = AUser.objects.get_or_create(
            username='admin_test',
            defaults={
                'email': 'admin_test@test.com',
                'user_type': AUser.UserType.ADMIN
            }
        )
        if created:
            admin_user.set_password('test_password')
            admin_user.save()
        
        excel_file = create_test_excel_student()
        excel_file.name = 'test_students.xlsx'
        
        client = Client()
        client.login(username='admin_test', password='test_password')
        
        response = client.post(
            '/admin/onboard/students',
            {
                'file': excel_file,
                'block_id': 'TEST_BLOCK_1'
            },
            follow=True
        )
        
        # Check if users were created
        student_users = AUser.objects.filter(email__startswith='s_')
        print(f"[STUDENT] Created {student_users.count()} student users")
        
        for user in student_users:
            profile = StudentProfile.objects.filter(user=user)
            status = "OK" if profile.exists() else "MISSING PROFILE"
            print(f"  - {user.email} (STU): {status}")
        
        if student_users.count() == 2:
            print("[PASS] Student upload successful")
            return True
        else:
            print("[FAIL] Expected 2 students, got " + str(student_users.count()))
            return False
            
    except Exception as e:
        print(f"[ERROR] Student test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Run all tests"""
    print("\n" + "="*60)
    print("BULK UPLOAD TEST SUITE")
    print("="*60)
    
    results = []
    
    results.append(("Examiner Upload", test_examiner_upload()))
    results.append(("Supervisor Upload", test_supervisor_upload()))
    results.append(("Student Upload", test_student_upload()))
    
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for name, result in results:
        status = "PASS" if result else "FAIL"
        print(f"{name}: {status}")
    
    print(f"\nTotal: {passed}/{total} tests passed")
    
    if passed == total:
        print("\nALL TESTS PASSED!")
        return True
    else:
        print(f"\n{total - passed} test(s) failed")
        return False


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
